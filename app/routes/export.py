from pathlib import Path
import shutil
import zipfile
from datetime import datetime

from fastapi import APIRouter, Request, Depends
from fastapi.responses import FileResponse, RedirectResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook

from app.database import get_db
from app.models import Trade

router = APIRouter()

EXPORT_DIR = Path("app/static/exports")
EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def require_login(request: Request):
    return bool(request.session.get("user_id"))


def clean_old_exports():
    for item in EXPORT_DIR.iterdir():
        try:
            if item.is_file():
                item.unlink()
            elif item.is_dir():
                shutil.rmtree(item)
        except Exception:
            pass


@router.get("/export")
def export_trades(request: Request, db: Session = Depends(get_db)):
    if not require_login(request):
        return RedirectResponse("/login", status_code=302)

    clean_old_exports()

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    export_folder = EXPORT_DIR / f"TradingJournal_Export_{timestamp}"
    screenshots_folder = export_folder / "screenshots"

    export_folder.mkdir(parents=True, exist_ok=True)
    screenshots_folder.mkdir(parents=True, exist_ok=True)

    excel_path = export_folder / "trades.xlsx"
    zip_path = EXPORT_DIR / f"TradingJournal_Export_{timestamp}.zip"

    trades = db.query(Trade).order_by(Trade.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"

    headers = [
        "ID", "Date", "Symbol", "Direction", "Entry Price", "Stop Loss",
        "Take Profit", "Lot Size", "Risk Amount", "Result", "Profit/Loss",
        "Strategy", "Emotion", "Mistake", "Notes",
        "Screenshot 1", "Screenshot 2", "Screenshot 3", "Screenshot 4", "Screenshot 5"
    ]

    ws.append(headers)

    for trade in trades:
        screenshot_links = []

        for index, image in enumerate(trade.images, start=1):
            source_path = Path("app" + image.image_path)

            if source_path.exists():
                file_name = f"trade_{trade.id}_chart_{index}{source_path.suffix}"
                destination_path = screenshots_folder / file_name

                shutil.copy2(source_path, destination_path)
                screenshot_links.append(f"screenshots/{file_name}")

        row = [
            trade.id,
            trade.trade_date,
            trade.symbol,
            trade.direction,
            trade.entry_price,
            trade.stop_loss,
            trade.take_profit,
            trade.lot_size,
            trade.risk_amount,
            trade.result,
            trade.profit_loss,
            trade.strategy,
            trade.emotion,
            trade.mistake,
            trade.notes,
        ]

        for i in range(5):
            row.append(f"Open Screenshot {i + 1}" if i < len(screenshot_links) else "")

        ws.append(row)

        current_row = ws.max_row

        for i, link in enumerate(screenshot_links[:5], start=16):
            cell = ws.cell(row=current_row, column=i)
            cell.hyperlink = link
            cell.style = "Hyperlink"

    wb.save(excel_path)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(excel_path, arcname="trades.xlsx")

        for file in screenshots_folder.iterdir():
            zipf.write(file, arcname=f"screenshots/{file.name}")

    return FileResponse(
        path=zip_path,
        filename=zip_path.name,
        media_type="application/zip"
    )